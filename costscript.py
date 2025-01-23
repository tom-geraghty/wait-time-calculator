#!/usr/bin/env python3

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def create_queueing_spreadsheet(filename="queueing_calculator.xlsx"):
    """
    Creates an Excel spreadsheet for a simplified Queueing Calculator with:
      - 2 header rows for INPUT/OUTPUT and Column Headings
      - 1 row of explanatory text
      - Rows for 'Current state' and 'Future state' data
    """

    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Queueing Calculator"

    # --- Define column content ---
    # We’ll have 12 columns in total:
    #  1) Scenario
    #  2) Your Utilisation
    #  3) Their Utilisation
    #  4) Requests per Week
    #  5) Base Wait Time (hrs) at Ref Util
    #  6) Ref Util
    #  7) Cost of Delay (£/hr)
    #  8) Willing to Pay (£/week)
    #  9) Calculated Wait Time (hrs/request)
    # 10) Total Delay (hrs/week)
    # 11) Total Delay Cost (£/week)
    # 12) Net Trade-Off (£/week)

    # Row 1: INPUT or OUTPUT
    row_types = [
        "INPUT",       # Scenario (label)
        "INPUT",       # Your Utilisation
        "INPUT",       # Their Utilisation
        "INPUT",       # Requests per Week
        "INPUT",       # Base Wait Time
        "INPUT",       # Ref Util
        "INPUT",       # Cost of Delay
        "INPUT",       # Willing to Pay
        "OUTPUT",      # Calculated Wait Time
        "OUTPUT",      # Total Delay
        "OUTPUT",      # Total Delay Cost
        "OUTPUT"       # Net Trade-Off
    ]

    # Row 2: Column Headings
    headers = [
        "Scenario",
        "Your Utilisation",
        "Their Utilisation",
        "Requests per Week",
        "Base Wait Time (hrs) at Ref Util",
        "Ref Util",
        "Cost of Delay (£/hr)",
        "Willing to Pay (£/week)",
        "Calculated Wait Time (hrs/request)",
        "Total Delay (hrs/week)",
        "Total Delay Cost (£/week)",
        "Net Trade-Off (£/week)"
    ]

    # Row 3: Explanatory text (where applicable)
    # (Leave blank if we don’t have a comment for a particular column)
    explanations = [
        "",  # Scenario
        "The fraction (or percentage) of your own time (or your team’s time) that is currently occupied by uninteruptible work.",
        "The fraction of another person’s or team’s time that is utilised.",
        "The number of times per week you need something (e.g., information, sign-off, etc.) from the other person/team. In queueing theory terms, this can be viewed as an arrival rate (λ).",
        "This is the baseline or reference wait time you have measured at some known 'Reference Utilisation'. For instance, you might have observed that when you (or they) were at 50% utilisation, you typically waited 2 hours for a response. That '2 hours' is your Base Wait Time.",
        "The utilisation level at which the Base Wait Time was measured. E.g., if you measured that 2-hour wait when utilisation was 50%, you’d put 0.50 here.",
        "The monetary cost (in £ or another currency) for each hour of delay. This can be a fully loaded labour cost or an estimate of revenue/opportunity cost lost per hour of delay. i.e. if you have a team of ten and pay roughly £20 per hour, it's £200 per hour (at least).",
        "How much extra you'd be willing to pay each week for higher utilisation, to weigh against the cost of delays. Basically, what do you think the value of this extra work that the team are doing is?",
        "The wait time per request, scaled from the base wait time by the current utilisation.",
        "The total weekly hours of delay, i.e. calculated wait time multiplied by the requests per week.",
        "The cost of that total delay, i.e. total delay hours multiplied by the cost of delay.",
        "Willing to Pay minus the Total Delay Cost, showing if you come out ahead or behind."
    ]

    # Write out row 1 (INPUT/OUTPUT labels)
    for col_idx, col_type in enumerate(row_types, start=1):
        ws.cell(row=1, column=col_idx, value=col_type)

    # Write out row 2 (Column Headings)
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=2, column=col_idx, value=header)

    # Write out row 3 (Explanations)
    for col_idx, expl in enumerate(explanations, start=1):
        ws.cell(row=3, column=col_idx, value=expl)

    # We’ll place scenario data in rows 4 (Current state) and 5 (Future state).

    # Example data for Current state
    current_state_data = [
        "Current state",  # Scenario
        0.50,             # Your Utilisation
        0.60,             # Their Utilisation (not used in formula)
        5,                # Requests per Week
        2.0,              # Base Wait Time
        0.50,             # Ref Util
        100.0,            # Cost of Delay
        0.0               # Willing to Pay
    ]

    # Example data for Future state
    future_state_data = [
        "Future state",
        0.90,   # Your Utilisation
        0.80,   # Their Utilisation
        5,
        2.0,
        0.50,
        100.0,
        500.0   # Willing to Pay
    ]

    # Put the data in row 4 and 5 for the input columns
    for row_idx, scenario_data in enumerate([current_state_data, future_state_data], start=4):
        for col_idx, val in enumerate(scenario_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Indices for columns (1-based)
    col_scenario         = 1
    col_rho_you          = 2
    col_rho_them         = 3  # currently unused in formula
    col_requests         = 4
    col_base_wait        = 5
    col_ref_util         = 6
    col_cost_of_delay    = 7
    col_willing_to_pay   = 8

    col_calc_wait        = 9  # Calculated Wait Time
    col_total_delay      = 10
    col_total_cost       = 11
    col_net_tradeoff     = 12

    # Apply formulas to rows 4 and 5
    for row in [4, 5]:
        cell_rho_you     = f"{get_column_letter(col_rho_you)}{row}"
        cell_requests    = f"{get_column_letter(col_requests)}{row}"
        cell_base_wait   = f"{get_column_letter(col_base_wait)}{row}"
        cell_ref_util    = f"{get_column_letter(col_ref_util)}{row}"
        cell_cost_delay  = f"{get_column_letter(col_cost_of_delay)}{row}"
        cell_wtp         = f"{get_column_letter(col_willing_to_pay)}{row}"

        # 1) Calculated Wait Time (hrs/request)
        #    Simple ratio approach: Wait = Base Wait * [(1 - RefUtil)/(1 - rho_you)] if (1 - rho_you) != 0
        #    else large number for near-infinite wait
        formula_wait = (
            f"=IF((1 - {cell_rho_you})=0,9999999,"
            f"{cell_base_wait}*((1 - {cell_ref_util})/(1 - {cell_rho_you})))"
        )
        ws.cell(row=row, column=col_calc_wait).value = formula_wait

        # 2) Total Delay (hrs/week) = per-request wait * requests/week
        cell_calc_wait = f"{get_column_letter(col_calc_wait)}{row}"
        formula_total_delay = f"={cell_calc_wait}*{cell_requests}"
        ws.cell(row=row, column=col_total_delay).value = formula_total_delay

        # 3) Total Delay Cost (£/week) = total delay * cost_of_delay
        cell_total_delay = f"{get_column_letter(col_total_delay)}{row}"
        formula_total_cost = f"={cell_total_delay}*{cell_cost_delay}"
        ws.cell(row=row, column=col_total_cost).value = formula_total_cost

        # 4) Net Trade-Off (£/week) = Willing to Pay - Total Delay Cost
        cell_total_cost = f"{get_column_letter(col_total_cost)}{row}"
        formula_tradeoff = f"={cell_wtp}-{cell_total_cost}"
        ws.cell(row=row, column=col_net_tradeoff).value = formula_tradeoff

    # Adjust column widths for readability
    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 35

    # Save the workbook
    wb.save(filename)
    print(f"Spreadsheet '{filename}' created successfully.")

if __name__ == "__main__":
    create_queueing_spreadsheet()
